"""
Generador de reporte Excel multi-hoja para Radar de Prensa CRTIC.

Uso:
    from app.excel_export import generate_excel_report
    excel_bytes = generate_excel_report(df, summary_data)
    # → pasar a st.download_button(data=excel_bytes, ...)
"""
import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
    numbers as xl_numbers,
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Paleta de estilos ──────────────────────────────────────────────────────────
_BLUE_DARK  = "1A237E"
_BLUE_MID   = "3949AB"
_BLUE_LIGHT = "E8EAF6"
_WHITE      = "FFFFFF"
_GRAY_ROW   = "F5F5F5"
_RED_LIGHT  = "FFCDD2"
_GREEN_LIGHT = "C8E6C9"
_YELLOW     = "FFF9C4"

_H_FONT  = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_H_FILL  = PatternFill("solid", fgColor=_BLUE_DARK)
_H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_SH_FONT  = Font(bold=True, color=_BLUE_DARK, size=10, name="Calibri")
_SH_FILL  = PatternFill("solid", fgColor=_BLUE_LIGHT)

_BODY_FONT  = Font(size=10, name="Calibri")
_WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
_CENTER     = Alignment(horizontal="center", vertical="top")
_LINK_FONT  = Font(color="1565C0", underline="single", size=10, name="Calibri")

_THIN = Side(border_style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CLP_FMT   = '$ #,##0'
_FLOAT_FMT = '0.00'

# Mapa de columnas de BD → nombre legible
_COL_MAP = {
    "fecha_deteccion":   "Fecha detección",
    "fecha_publicacion": "Fecha publicación",
    "medio":             "Medio",
    "titulo":            "Título",
    "url":               "URL",
    "keyword":           "Keyword",
    "tipo_mencion":      "Tipo de mención",
    "sentimiento":       "Sentimiento",
    "relevancia":        "Relevancia",
    "area_crtic":        "Área CRTIC",
    "valor_base_medio":  "Valor base (CLP)",
    "factor_visibilidad":"Factor visibilidad",
    "factor_estrategico":"Factor estratégico",
    "vem":               "VEM referencial (CLP)",
    "estado":            "Estado",
}

_ORDERED_COLS = list(_COL_MAP.keys())
_CLP_COLS     = {"valor_base_medio", "vem"}
_FLOAT_COLS   = {"factor_visibilidad", "factor_estrategico"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _set_header(ws, row: int, col: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = _H_FONT
    c.fill   = _H_FILL
    c.alignment = _H_ALIGN
    c.border = _BORDER
    return c


def _set_subheader(ws, row: int, col: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = _SH_FONT
    c.fill   = _SH_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _BORDER
    return c


def _set_value(ws, row: int, col: int, value, fmt=None, link=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _BODY_FONT
    c.alignment = _WRAP_ALIGN
    c.border    = _BORDER
    if fmt:
        c.number_format = fmt
    if link and isinstance(link, str) and link.startswith("http"):
        c.hyperlink = link
        c.font = _LINK_FONT
    return c


def _autowidth(ws, min_w=10, max_w=60):
    """Ajusta el ancho de cada columna al contenido."""
    for col_cells in ws.columns:
        width = min_w
        for cell in col_cells:
            try:
                v = str(cell.value or "")
                # No contar saltos de línea para calcular ancho
                longest_line = max((len(line) for line in v.split("\n")), default=0)
                width = max(width, longest_line + 2)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width, max_w)


def _shade_rows(ws, start_row: int, end_row: int, n_cols: int):
    """Filas alternas sombreadas."""
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            fill = PatternFill("solid", fgColor=_GRAY_ROW)
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill


def _write_df_sheet(ws, df: pd.DataFrame, title: str = ""):
    """
    Escribe un DataFrame en una hoja con encabezados formateados,
    filtros, filas alternas y formatos de número.
    """
    # Seleccionar y renombrar columnas disponibles
    avail = [c for c in _ORDERED_COLS if c in df.columns]
    sub   = df[avail].copy()
    sub.rename(columns={k: _COL_MAP[k] for k in avail}, inplace=True)

    # Fila de título opcional
    offset = 0
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(avail))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(bold=True, size=13, color=_BLUE_DARK, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
        ws.row_dimensions[1].height = 24
        offset = 1

    header_row = 1 + offset
    for ci, col_name in enumerate(sub.columns, start=1):
        _set_header(ws, header_row, ci, col_name)
    ws.row_dimensions[header_row].height = 28

    # Datos
    url_col_idx = None
    if "URL" in sub.columns:
        url_col_idx = list(sub.columns).index("URL") + 1

    for ri, row in enumerate(sub.itertuples(index=False), start=header_row + 1):
        for ci, (col_name, val) in enumerate(zip(sub.columns, row), start=1):
            orig_col = avail[ci - 1]
            fmt  = _CLP_FMT   if orig_col in _CLP_COLS else (
                   _FLOAT_FMT if orig_col in _FLOAT_COLS else None)
            link = val if ci == url_col_idx and isinstance(val, str) else None
            # Valor limpio para URL (texto acortado)
            display = val
            if link:
                try:
                    from urllib.parse import urlparse
                    display = urlparse(val).netloc or val
                except Exception:
                    display = val
            _set_value(ws, ri, ci, display if link else val, fmt=fmt, link=link)

    data_end = header_row + len(sub)

    # Filtros
    if len(sub) > 0:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(avail))}{data_end}"
        )

    # Filas alternas
    _shade_rows(ws, header_row + 1, data_end, len(avail))

    # Congelar fila de encabezado
    ws.freeze_panes = f"A{header_row + 1}"

    _autowidth(ws)
    return sub


# ── Hoja Resumen ───────────────────────────────────────────────────────────────

def _build_resumen(ws, df: pd.DataFrame, summary: dict):
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24

    rows = [
        ("RADAR DE PRENSA CRTIC — RESUMEN", None, True),
        (None, None, False),
        ("Generado el", summary.get("generado", datetime.now().strftime("%d/%m/%Y %H:%M")), False),
        ("Período", summary.get("periodo", ""), False),
        (None, None, False),
        ("MÉTRICAS GENERALES", None, "section"),
        ("Total de menciones", len(df), False),
        ("VEM total referencial (CLP)", df["vem"].sum() if "vem" in df.columns else 0, "clp"),
        (None, None, False),
        ("MENCIONES POR SENTIMIENTO", None, "section"),
        ("Positivas",
         int((df["sentimiento"] == "Positivo").sum()) if "sentimiento" in df.columns else 0, False),
        ("Neutras",
         int((df["sentimiento"] == "Neutro").sum())   if "sentimiento" in df.columns else 0, False),
        ("Negativas",
         int((df["sentimiento"] == "Negativo").sum()) if "sentimiento" in df.columns else 0, False),
        (None, None, False),
        ("MENCIONES POR RELEVANCIA", None, "section"),
        ("Alta",
         int((df["relevancia"] == "Alta").sum())  if "relevancia" in df.columns else 0, False),
        ("Media",
         int((df["relevancia"] == "Media").sum()) if "relevancia" in df.columns else 0, False),
        ("Baja",
         int((df["relevancia"] == "Baja").sum())  if "relevancia" in df.columns else 0, False),
        (None, None, False),
        ("TOP 5 MEDIOS POR VEM", None, "section"),
    ]

    r = 1
    for label, value, kind in rows:
        if kind is True:          # Título principal
            ws.merge_cells(f"A{r}:B{r}")
            c = ws.cell(row=r, column=1, value=label)
            c.font  = Font(bold=True, size=14, color=_WHITE, name="Calibri")
            c.fill  = PatternFill("solid", fgColor=_BLUE_DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 30
        elif kind == "section":   # Sub-encabezado de sección
            ws.merge_cells(f"A{r}:B{r}")
            _set_subheader(ws, r, 1, label)
            ws.cell(row=r, column=2)  # merged
        elif label is None:       # Fila vacía
            pass
        else:
            ca = ws.cell(row=r, column=1, value=label)
            ca.font = Font(bold=True, size=10, name="Calibri")
            ca.border = _BORDER
            ca.alignment = Alignment(vertical="center")
            cb = ws.cell(row=r, column=2, value=value)
            cb.font   = _BODY_FONT
            cb.border = _BORDER
            cb.alignment = Alignment(horizontal="right", vertical="center")
            if kind == "clp":
                cb.number_format = _CLP_FMT
        r += 1

    # Top 5 medios por VEM
    if "medio" in df.columns and "vem" in df.columns:
        top5 = df.groupby("medio")["vem"].sum().nlargest(5).reset_index()
        _set_header(ws, r, 1, "Medio")
        _set_header(ws, r, 2, "VEM referencial (CLP)")
        r += 1
        for _, row in top5.iterrows():
            ws.cell(row=r, column=1, value=row["medio"]).border = _BORDER
            c2 = ws.cell(row=r, column=2, value=row["vem"])
            c2.number_format = _CLP_FMT
            c2.border = _BORDER
            c2.alignment = Alignment(horizontal="right")
            r += 1

    ws.freeze_panes = "A2"


# ── Hoja Alertas ───────────────────────────────────────────────────────────────

def _build_alertas(ws, df: pd.DataFrame):
    if df.empty:
        ws.cell(row=1, column=1, value="Sin alertas para el período seleccionado.")
        return

    # Negativas
    neg = df[df["sentimiento"] == "Negativo"] if "sentimiento" in df.columns else pd.DataFrame()
    # Alta relevancia
    alta = df[df["relevancia"] == "Alta"] if "relevancia" in df.columns else pd.DataFrame()
    # Posible error de denominación
    def _es_error_denom(row):
        text = f"{row.get('titulo', '')} {row.get('snippet', '')}".lower()
        return "crtic" not in text and "tecnocreat" not in text
    error_denom = df[df.apply(_es_error_denom, axis=1)] if not df.empty else pd.DataFrame()

    r = 1
    sections = [
        ("⚠️ Menciones negativas", neg, _RED_LIGHT),
        ("⭐ Menciones de alta relevancia", alta, _YELLOW),
        ("📝 Posible error de denominación", error_denom, _GREEN_LIGHT),
    ]

    avail = [c for c in _ORDERED_COLS if c in df.columns]

    for section_title, section_df, row_color in sections:
        # Título de sección
        ws.merge_cells(
            start_row=r, start_column=1, end_row=r, end_column=len(avail)
        )
        c = ws.cell(row=r, column=1, value=section_title)
        c.font  = Font(bold=True, size=12, color=_BLUE_DARK, name="Calibri")
        c.fill  = PatternFill("solid", fgColor=_BLUE_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        if section_df.empty:
            ws.cell(row=r, column=1, value="(Sin registros)").font = Font(italic=True, color="9E9E9E")
            r += 2
            continue

        # Encabezados
        for ci, col_key in enumerate(avail, start=1):
            _set_header(ws, r, ci, _COL_MAP[col_key])
        r += 1

        # Filas
        url_col_idx = avail.index("url") + 1 if "url" in avail else None
        for _, row_data in section_df[avail].iterrows():
            fill = PatternFill("solid", fgColor=row_color)
            for ci, col_key in enumerate(avail, start=1):
                val = row_data[col_key]
                fmt  = _CLP_FMT   if col_key in _CLP_COLS else (
                       _FLOAT_FMT if col_key in _FLOAT_COLS else None)
                link = val if ci == url_col_idx and isinstance(val, str) else None
                display = val
                if link:
                    try:
                        from urllib.parse import urlparse
                        display = urlparse(val).netloc or val
                    except Exception:
                        display = val
                cell = _set_value(ws, r, ci, display if link else val, fmt=fmt, link=link)
                cell.fill = fill
            r += 1
        r += 1  # Separador entre secciones

    ws.auto_filter.ref = None  # Sin filtro global en alertas (secciones mixtas)
    ws.freeze_panes = "A2"
    _autowidth(ws)


# ── Función principal ──────────────────────────────────────────────────────────

def generate_excel_report(df: pd.DataFrame, summary_data: dict) -> bytes:
    """
    Genera un reporte Excel multi-hoja como bytes.

    Args:
        df:           DataFrame con todas las menciones del período.
        summary_data: dict con claves:
                        'periodo'  → str, ej. "01/04/2026 – 30/04/2026"
                        'generado' → str, ej. "30/04/2026 09:00"

    Returns:
        bytes del archivo .xlsx, listo para st.download_button.
    """
    wb = Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen"
    _build_resumen(ws_res, df, summary_data)

    # ── Hoja 2: Menciones ─────────────────────────────────────────
    ws_men = wb.create_sheet("Menciones")
    _write_df_sheet(ws_men, df, title="Menciones del período")

    # ── Hoja 3: Top menciones por VEM ────────────────────────────
    ws_top = wb.create_sheet("Top menciones por VEM")
    if not df.empty and "vem" in df.columns:
        df_top = df.sort_values("vem", ascending=False).head(50)
    else:
        df_top = df
    _write_df_sheet(ws_top, df_top, title="Top menciones por VEM referencial")

    # ── Hoja 4: Alertas ───────────────────────────────────────────
    ws_alert = wb.create_sheet("Alertas")
    _build_alertas(ws_alert, df)

    # ── Serializar ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
